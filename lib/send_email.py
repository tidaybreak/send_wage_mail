#!/usr/bin/env python
# -*-coding:utf-8-*-

import os
import sys
import codecs
import time
import datetime
import smtplib
import shutil
import ConfigParser
import base64 
from email.mime.text import MIMEText
from email.header import Header
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import  PatternFill, Font, Fill, Border, Side
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from openpyxl.styles import Alignment

current_dir = os.path.dirname(os.path.abspath(__file__))
log_path = current_dir + os.sep + 'log.txt'


def loginfo(msg):
    with codecs.open(log_path, 'a', 'utf-8') as f:
        f.write(time.strftime("%Y-%m-%d %X") + "-" + msg.decode('utf-8') + os.linesep)


def send_mail(to_addr, subject, html_template, filename, file, user_mail, user_passwd, smtp_server, smtp_port, enable_ssl):
    try:
        message = MIMEMultipart()
        message['From'] = Header(user_mail)
        message['To'] = Header(to_addr)
        message['Subject'] = Header(subject, 'utf-8')
		
        #---这是文字部分---
        part = MIMEText(html_template, 'html', 'utf-8')
        message.attach(part)
		
        part = MIMEApplication(open(file,'rb').read())
        #part.add_header('Content-Disposition', 'attachment', filename=filename)
        part.add_header('Content-Disposition', 'attachment', filename= '=?utf-8?b?' + base64.b64encode(filename.encode('UTF-8')) + '?=') 
        message.attach(part)

        mail_obj = None
        if enable_ssl:
            mail_obj = smtplib.SMTP_SSL(smtp_server, smtp_port)
        else:
            mail_obj = smtplib.SMTP(smtp_server, smtp_port)
        mail_obj.login(user_mail, user_passwd)
        mail_obj.sendmail(user_mail, to_addr, message.as_string())
        mail_obj.quit()
        return True
    except Exception as e:
        loginfo('send mail to ' + str(to_addr) + ' failed,exception: ' + str(e))
        return False


def read_data(excel_file):
    mail_title = ''
    mail_content = ''
    table_title = ''
    data = []
    titles = []
    # rows number one staff in excel
    staff_rows = []
    wb = load_workbook(filename=excel_file, read_only=False, data_only=True)
    ws = wb.worksheets[0]
    first_line = True
    idx = 0
    for row in ws.rows:
        idx += 1
        if idx == 2:
            mail_title = row[1].value
        if idx == 3:
            mail_content = row[1].value
        if idx == 4:
            table_title = row[1].value
        if idx <= 4:
            continue

        item = []
        first_column = True
        for cell in row:
            if first_line:
                titles.append(cell.value)
            else:
                if first_column:
                    rows_check = check_merge(cell.row, cell.col_idx, ws.merged_cells)
                    if rows_check["type"] == 'rowspan':
                        staff_rows.append(rows_check["rowspan"])
                    elif rows_check["type"] == 'normal':
                        staff_rows.append(1)
                item.append({
                    "value": cell.value,
                    "coordinate": cell.coordinate,
                    "col": cell.col_idx,
                    "row": cell.row
                })
            first_column = False
        if not first_line:
            data.append(item)
        first_line = False
    return mail_title, mail_content, table_title, titles, data, ws.merged_cells, staff_rows


def finish(excel_file, row):
    wb = load_workbook(filename=excel_file, read_only=False, data_only=True)
    ws = wb.worksheets[0]
    ws.cell(row=row, column=1).value = 'ok'
    wb.save(filename= excel_file)
    return 
	
def check_merge(row, col, merged_cells):
    for item in merged_cells.ranges:
        # on the same column
        if item.min_col == item.max_col == col:
            # rowspan
            if item.min_row == row:
                return {"type": "rowspan", "rowspan": item.max_row - item.min_row + 1}
            elif item.min_row < row <= item.max_row:
                return {"type": "none"}
        # on the same row
        elif item.max_row == item.min_row == row:
            # colspan
            if item.min_col == col:
                return {"type": "colspan", "colspan": item.max_col - item.min_col + 1}
            elif item.min_col < col <= item.max_col:
                return {"type": "none"}
        elif item.min_row == row and item.min_col == col:
            return {"type": "mix", "rowspan": item.max_row - item.min_row + 1,
                    "colspan": item.max_col - item.min_col + 1}
        elif item.min_row <= row <= item.max_row and item.min_col <= col <= item.max_col:
            return {"type": "none"}
    return {"type": "normal"}

def is_writeable(excel_file, check_parent=False):
    wb = load_workbook(filename=excel_file, read_only=False, data_only=True)
    try:
        wb.save(filename= excel_file)
        return True
    except Exception as e:
        return False
    

def main():
    send_table = True
    if len(sys.argv) == 2:
        send_table = False
    cf = ConfigParser.ConfigParser()
    cf.read(current_dir + os.sep + 'config.ini')
    user = cf.get('user', 'email')
    user = user.replace("\r","")
    user = user.replace("\n","")
    pwd = cf.get('user', 'password')
    pwd = pwd.replace("\r","")
    pwd = pwd.replace("\n","")
    server = cf.get('user', 'smtp_server')
    server = server.replace("\r","")
    server = server.replace("\n","")
    port = cf.getint('user', 'smtp_port')
    enable_ssl = cf.getboolean('user', 'enable_ssl')
	
    try:
        shutil.rmtree('../个人详情'.decode('utf-8').encode('gbk'))  
        os.mkdir('../个人详情'.decode('utf-8').encode('gbk'))
    except Exception as e:
        print '请先关闭"个人详情"文件夹中的文件!'.decode('utf-8').encode('gbk')
        return False
	
    money = '工资表'
    money = money.decode('utf-8').encode('gbk')
    money_file = current_dir + os.sep + '..\\' + money + '.xlsx'
	
	
    mail_title, mail_content, table_title, titles, salary_data, merged_cells, staff_rows = read_data(money_file)
    html_template = '<table border="0" style="border-collapse:collapse">'
    html_template += '<tr border="0"><td border="0" colspan="22" style="font-size:16px;padding-top:20px;padding-bottom:20px;">%s' % mail_content
    html_template += '</td></tr>'
    html_template += '</table>'
	
    if send_table == True:
        html_template += '<table border="1" style="border-collapse:collapse">'
        html_template += '<thead>'
        html_template += '<tr><th colspan="22" style="font-size:20px;padding-top:20px;padding-bottom:20px;">%s' % table_title
        html_template += '</th></tr>'
        html_template += '<tr>'
        titles = ['' if v is None else v for v in titles]
        for title in titles[2:]:
            html_template += '<th style="padding-left:20px;padding-right:20px">' + title + '</th>'
        html_template += '</tr>'
        html_template += '</thead>'
        html_template += '<tbody>'
        html_template += '<<placeholder>>'
        html_template += '</tbody>'
        html_template += '</table>'

    today_day = datetime.datetime.now().day
    today_month = datetime.datetime.now().month
    print 'The Company paid wages before the 5th'
    print 'Today is ' + time.strftime("%B %d")
    #mail_subject = '%s月份工资条，请查收'
    mail_subject = mail_title
    # Pay money before the 5th of each month
    #if today_day > 31:
    #    mail_subject = mail_subject % today_month
    #else:
    #    today_month = today_month - 1
    #    if today_month == 0:
    #        today_month = 12
    #    mail_subject = mail_subject % today_month
    english_month = datetime.date(1900, today_month, 1).strftime('%B')
    print 'The mail subject will be show as "' + english_month + ' salley bill"'
    print "\n"
    has_failture = False
    row_index = 0
    for staff_row in staff_rows:
        if is_writeable(money_file) == False:
            print '请先关闭工资表.xlsx'.decode('utf-8').encode('gbk')
            break
        staff_status = salary_data[row_index][0]["value"]
        staff_email = salary_data[row_index][1]["value"]
        holder_str = ''
        for item in salary_data[row_index:row_index + staff_row]:
            holder_str += '<tr>'
            idx = 0
            for i in item[2:]:
                idx += 1
                width = 200
                if idx == 1:
                    width = 200
                check = check_merge(i["row"], i["col"], merged_cells)
                try:
                    val = '' if i["value"] is None else i["value"]
                except Exception as e:
                    print e
                if check["type"] == 'rowspan':
                    holder_str += '<td width="%s" style="padding-left:5px;padding-right:5px;" rowspan="%s">%s</td>' % (
                        width, check["rowspan"], val)
                if check["type"] == 'colspan':
                    holder_str += '<td width="%s" style="text-align:center;" colspan="%s">%s</td>' % (width, check["colspan"], val)
                if check["type"] == 'mix':
                    holder_str += '<td width="%s" style="text-align:center;" rowspan="%s" colspan="%s">%s</td>' % (
                    width, check["rowspan"], check["colspan"], val)
                if check["type"] == 'none':
                    pass
                if check["type"] == 'normal':
                    holder_str += '<td width="%d" style="padding-left:5px;padding-right:5px;">%s</td>' % (width, val)
            holder_str += '</tr>'
        if send_table == True:
            html_content = html_template.replace('<<placeholder>>', holder_str)
        else:
            html_content = html_template
			
        staff_idx = salary_data[row_index][2]["value"]
        if staff_idx is None:
            staff_idx = ''
        staff_idx = str(staff_idx)
        staff_name = salary_data[row_index][3]["value"]
		
        if staff_name is not None and staff_email is not None:
            wb = Workbook()
            ws = wb.get_active_sheet()
            ws.title = 'money'  # 设置worksheet的标题
            ws.merge_cells('A1:V1')
            ws.merge_cells('A2:V2')
            ws.row_dimensions[1].height = 40
            ws.row_dimensions[2].height = 40
            ws.row_dimensions[3].height = 48
            ws.cell(row=1, column=1).value = '广东奥飞数据科技股份有限公司'
            ws.cell(row=2, column=1).value = table_title
            align = Alignment(horizontal='center',vertical='center',wrap_text=True)
            ws.cell(row=1, column=1).alignment = align
            ws.cell(row=2, column=1).alignment = align
			
            left, right, top, bottom = [Side(style='thin',color='000000')]*4
            border = Border(left=left, right=right, top=top, bottom=bottom)

            column = 0
            fill = PatternFill("solid", fgColor="5B9BD5")
            font = Font(color="FFFFFF")
            for title in titles[2:]:
                column += 1
                ws.cell(row=3, column=column).value = title
                ws.cell(row=3, column=column).border = border
                #ws.cell(row=3, column=column).fill = fill
                #ws.cell(row=3, column=column).font = font
                ws.cell(row=3, column=column).alignment = align
            column += 1
            ws.cell(row=3, column=column).value = '邮箱'
            ws.cell(row=3, column=column).alignment = align
            ws.cell(row=3, column=column).border = border
            row = 3
            for item in salary_data[row_index:row_index + staff_row]:
                row += 1
                idx = 0
                for i in item[2:]:
                    idx += 1
                    width = 200
                    check = check_merge(i["row"], i["col"], merged_cells)
                    try:
                        val = '' if i["value"] is None else i["value"]
                    except Exception as e:
                        print e
                    ws.cell(row=row, column=idx).value = val
                    ws.cell(row=row, column=idx).alignment = align
                    ws.cell(row=row, column=idx).border = border
            idx += 1
            ws.cell(row=row, column=idx).value = staff_email
            ws.cell(row=row, column=idx).alignment = align
            ws.cell(row=row, column=idx).border = border
            wb.save(filename= u'../个人详情/' + staff_name + u'工资条' + staff_idx + '.xlsx')
								
        if staff_email is not None and staff_status != 'ok':
        #if staff_email is not None:
            #time.sleep(5)
            send_result = send_mail(staff_email, mail_subject, html_content,  staff_name + u'工资条' + staff_idx + '.xlsx',  u'../个人详情/' + staff_name + u'工资条' + staff_idx + '.xlsx', user, pwd, server, port, enable_ssl)
            if not send_result:
                has_failture = True
                print 'mail to:' + str(staff_email) + ' failed!!!,please send this email manually.'
                loginfo('mail to:' + str(staff_email) + ' failed!!!,please send this email manually.')
            else:
                finish(money_file, row_index + 6)
                print 'mail to:' + str(staff_email) + ' Successfully'
                time.sleep(1)


        row_index += staff_row
    print "\n"
    if has_failture:
        print "There are some mails failed to be send, please check theme in the log.txt"
        print "\n"
        raw_input('Please input any key to quit...')
    else:
        print "Program has run successfully,all the mails have been sent successfully."
        print 'The program will exit in 3 seconds...'
        time.sleep(3)
    sys.exit(0)


main()
